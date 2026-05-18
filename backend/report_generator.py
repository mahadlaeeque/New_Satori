"""
Report Generator for Satori AI.
Uses Gemini to design TABULAR report structure, BigQuery to fetch data,
then generates clean, professional Excel/PDF files with data tables and totals.
"""
import os, json, io, datetime
from google import genai
from bigquery_client import run_query, get_table_schema, discover_tables
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable


# ── Brand colors ──
BRAND_GREEN = "8AC441"
BRAND_DARK = "333333"
BRAND_PURPLE = "353085"
BRAND_TEAL = "0A5F89"


# ── Shared SAP schema reference ──
# Used by every report-related Gemini prompt so the AI can plan JOINs and be
# honest about what data is and isn't available. Kept in sync with
# `_DASHBOARD_SAP_SCHEMAS` in main.py — both describe the same mirror.
_REPORT_SAP_SCHEMAS = """Detailed table schemas (BigQuery project `sfml-491907`, dataset `sap_hana_mirror`). Column TYPES are shown in parentheses — use them correctly.

PRIMARY DW REPORTING TABLES (preferred for stock + movement reports):
- fact_material_stock_daily — daily OPENING/CLOSING BALANCE. **3-row-per-key model**: stock_type (STRING — 'STORAGE' / 'SPECIAL' / 'RATE'), plant (STRING), storage_location (STRING — only on STORAGE rows; NULL for SPECIAL/RATE), material_id (STRING), material_type (STRING), base_unit_of_measure (STRING), posting_date (DATE), cumulative_qty (NUMERIC). Per-material qty = SUM(STORAGE.cumulative_qty)+SUM(SPECIAL.cumulative_qty); rate = MAX(RATE.cumulative_qty); value = qty × rate. NO `stock_value` / `stock_qty` columns. Plants populated: 1100/2100/3100. Date range 2025-04-15 → current.
- fact_material_movements_daily — daily Receipts/Issues/Adjustments per (plant, storage_location, material_id): posting_date (DATE), plant, storage_location, material_id, material_type, base_unit_of_measure, RECEIPT_QTY, RECEIPT_VALUE, ISSUE_QTY, ISSUE_VALUE, ADJUSTMENT_QTY, ADJUSTMENT_VALUE (all NUMERIC). ISSUE/ADJUSTMENT are signed (issues mostly negative; adjustments either sign). Use plain `SUM(...)` and present with natural sign — only flip on explicit "absolute" / "magnitude" requests. NEVER `SUM(ABS(...))` (double-counts reversals).

DIMENSION + REFERENCE TABLES:
- plants (9 rows): plant_id, plant_name, plant_name_2, city, country_code, region (state code), purchasing_org, sales_org, distribution_channel, division, valuation_area (all STRING). Active plant_ids: '1000' (Corporate Office), '1100' (Fabric Processing), '1101' (Bio Mass), '2100' (Work Wear), '3100' (Fashion Wear) — all Lahore, PK.
- material_master (~570K rows): material_id (STRING), material_type, base_unit_of_measure, creation_date (STRING YYYYMMDD), last_change_date (STRING YYYYMMDD).
- material_descriptions (~570K rows): material_id, language_key ('E'=English, 'N'=Native), material_description (all STRING). For readable names JOIN ON material_id AND language_key='E'.
- material_valuation (1.68M rows): material_id, valuation_area, valuation_type, moving_average_price (NUMERIC, ~5% populated), standard_price (NUMERIC, ~0.4% populated), price_control_indicator, price_unit, valuation_class, fiscal_year, fiscal_period. WARNING: total_stock_quantity / total_stock_value are ZERO — DO NOT use them; derive from fact_material_stock_daily as qty × rate.

OPERATIONAL TABLES:
- material_documents (23.6M rows): raw movement events. Each event = TWO rows (debit_credit_indicator 'S' + 'H') — filter one side when counting events. Cols: material_doc_number, material_id, plant, storage_location, batch, quantity (NUMERIC), amount_local_currency (NUMERIC), stock_qty, stock_value, posting_date (STRING YYYYMMDD), entry_date, document_date, purchase_order, order_number, unit_of_measure, movement_type (101 GR / 261 GI / 311 transfer / 601 delivery / Z87 custom / etc.).
- orders (506K rows): order_type (STRING — Z306/Z302/PMWO/CMWO/…), client_id, order_number, creation_date, change_date (STRING YYYYMMDD). NO plant column — derive plant via material_documents.order_number.
- purchase_order_header (49K rows): purchase_order_type (Z001/Z002/Z003/NB/ZAN/…), client_id, purchase_order_number, purchase_order_date (STRING YYYYMMDD), change_date. Header-only — no plant or amount; JOIN accounting_doc_segment ON purchase_order_number for those.
- accounting_doc_segment (29.9M rows): client_id, company_code ('1000'), belnr, gjahr (STRING '2023'-'2026'), buzei, purchase_order_number, material_id, plant_id, posting_key, debit_credit_indicator ('S'=debit, 'H'=credit), amount_local_currency (NUMERIC, PKR), tax_code, posting_date (STRING YYYYMMDD), document_type (ML/WA/WL/WE/RE/SA/KR/KZ/RV/DR/…).
- universal_journal (29.3M rows): client_id, company_code ('1000'), fiscal_year (STRING '2025'/'2026'), doc_number, doc_line, account, transaction_currency (PKR primary), fiscal_period (STRING '000'-'012'), posting_key, posting_date (STRING YYYYMMDD), material_id, plant_id, amount_local_currency, reference_document.

JOIN HINTS — these are the joins you can rely on:
- `<any>.material_id` ↔ `material_descriptions.material_id` (with language_key='E') for readable material name.
- `<any>.material_id` ↔ `material_master.material_id` for material_type / base_unit_of_measure.
- `<any>.plant`/`plant_id` ↔ `plants.plant_id` for plant_name, city, region.
- `material_documents.purchase_order` ↔ `purchase_order_header.purchase_order_number` ↔ `accounting_doc_segment.purchase_order_number` for PO chain.
- `accounting_doc_segment.belnr + gjahr` ↔ `universal_journal.doc_number + fiscal_year` for FI ↔ universal-journal links.
- MATERIAL ID LEADING-ZERO RULE: stored material_id is 18-char zero-padded. Match user input via `LTRIM(material_id,'0') = LTRIM(:input,'0')` so trimmed forms work too.

DATE RULES:
- DW fact tables (`fact_material_stock_daily`, `fact_material_movements_daily`) `posting_date` is DATE — compare directly.
- Other SAP date columns (posting_date in material_documents/accounting_doc_segment/universal_journal, purchase_order_date, creation_date, change_date, document_date) are STRING YYYYMMDD — wrap with `SAFE.PARSE_DATE('%Y%m%d', col)` for range filters.
- fiscal_year, fiscal_period, gjahr are STRING.

WHAT IS NOT IN THIS MIRROR:
- No customer-facing sales orders, customer master, dealer / distributor data, or revenue-by-customer tables. The closest proxies are document_type='RV'/'DR' in accounting_doc_segment (billing / customer-invoice postings) — label these as "SAP-side proxy", not true revenue.
- No HR / employee / region-head / org-chart tables. If a user asks for a person's name or role, you cannot join to it — say so plainly and offer the closest available alternative (e.g., the plant name, or the user_id field on the row if present).
- No purchase_order_item / line-quantity / open-PO-quantity table — purchase_order_header has only header dates and types.
- material_valuation.total_stock_quantity and total_stock_value are ZERO across all rows — never query them."""


def get_ai_client():
    """Get Gemini client."""
    project = os.environ.get("VERTEX_PROJECT", "")
    location = os.environ.get("VERTEX_LOCATION", "us-central1")
    api_key = os.environ.get("GEMINI_API_KEY", "")
    if project:
        return genai.Client(vertexai=True, project=project, location=location)
    if api_key:
        return genai.Client(api_key=api_key)
    raise Exception("No AI backend configured")


# ── AI Prompt for structured Q&A (used by /api/report/refine) ──

REPORT_REFINE_PROMPT = """You are Satori AI, a smart business analyst at a company running SAP ERP. You help users generate tabular reports from their enterprise data.

You know the company's data inside out. When a user asks for a report, you ALREADY KNOW which table to use, which columns make sense, and how to structure the report. You use your intelligence — don't ask the user technical questions about column names, table names, or SQL.

AVAILABLE DATA (use this knowledge internally — never expose raw column names to the user):
{tables}

HOW TO BEHAVE:

1. FIRST MESSAGE — Acknowledge the request warmly. Then ask 1-2 simple BUSINESS questions to understand their needs. You MUST end your first message with a question. Do NOT present a summary yet. Examples:
   - "Sure, I can pull a stock report! Quick questions — do you want current on-hand stock across all plants, or zoomed into one plant? And should it be by material, by storage location, or both?"
   - "Happy to put together a purchase-order report. Should I cover all PO types or a specific one, and any particular date range?"
   - "Great, I'll prepare a financial summary from the universal journal. Do you want it broken down by fiscal period, by plant, or both?"
   NEVER list raw column names. NEVER present the summary on the first message. Just ask questions.

2. SECOND MESSAGE (after user answers) — NOW present a clean summary of what you'll generate and ask for confirmation:

   Here's what I'll generate:

   **Stock On Hand — Current Snapshot**
   Shows plant, storage location, material ID, material description, stock quantity, stock value
   Grouped by plant and storage location, sorted by stock value (largest first)
   Includes total stock value row
   Format: Excel

   Say **"generate"** to create it, or tell me what to change.

   Keep the summary in BUSINESS language. Never mention SQL, table names, WHERE clauses, or column names with underscores.

3. THIRD MESSAGE (when user confirms with "generate", "go", "yes", "looks good", etc.) — Return EXACTLY this JSON and nothing else:
   {{"ready": true, "config": {{"title": "Report Title", "table": "BigQueryTableName", "columns": ["col1", "col2"], "filters": "SQL WHERE clause or empty string", "group_by": "SQL GROUP BY clause or empty string", "order_by": "SQL ORDER BY clause or empty string", "aggregations": {{"colName": "SUM"}}, "format": "excel"}}}}

CRITICAL RULES:
- NEVER expose technical details (table names, column names, SQL) to the user. Translate everything into business language.
- NEVER output the JSON until the user explicitly confirms. The JSON is only for step 3.
- YOU decide the best columns, grouping, and sort order based on what makes business sense. Only ask the user about business preferences (time period, specific customer/vendor, level of detail).
- Default to Excel format unless the user asks for PDF.
- Be concise — max 3-4 sentences per message.
"""


# Edit-mode system prompt — used when the user is modifying an existing,
# already-saved report. The AI is shown the FULL current config (including the
# generated SQL and the columns the query produces) plus the SAP schema with
# join hints, so it can decide whether a requested change is a simple
# column-add or needs a JOIN to another table — and be honest when the data
# doesn't exist in this mirror at all.
REPORT_EDIT_PROMPT = """You are Satori AI, a smart business analyst at a company running SAP ERP. You help users EDIT their existing tabular reports.

THE REPORT THE USER IS EDITING (this is your source of truth — don't re-ask things already answered here, and don't show this raw to the user — translate into business language):

{current_config}

Pay special attention to:
- `sql` — the BigQuery SQL currently powering this report. Read it before deciding what's possible. Adding a column means adapting THIS query; you don't get to start over silently.
- `all_columns` — every column the SQL currently returns.
- `columns` — the user-visible subset (a removed column may still be selectable via the UI's add-column menu, no AI needed).
- `filters`, `group_by`, `order_by` — the current scope. Don't re-ask the user about date ranges or plant scope when these are already populated.

""" + _REPORT_SAP_SCHEMAS + """

WHEN THE USER ASKS TO ADD A COLUMN — follow this decision tree (silently, in your head):
1. Is the column already in `all_columns`? → tell the user it's already in the report (or suggest re-adding it from the UI's "Add column" menu — no AI work needed).
2. Is the column on the SAME base table the report already queries? → just add it to the SELECT list of the existing SQL.
3. Is the column on a DIFFERENT table that has a known JOIN (see "JOIN HINTS" above)? → add a LEFT JOIN to the existing SQL. Use a LEFT (not INNER) join so existing rows aren't dropped if the lookup misses. Examples:
   - "Material name / description" → LEFT JOIN material_descriptions ON material_id AND language_key='E'.
   - "Plant name" / "city" / "region (state code)" → LEFT JOIN plants ON plant_id.
   - "Material type" / "unit of measure" → LEFT JOIN material_master ON material_id.
   - "PO type" / "PO date" → LEFT JOIN purchase_order_header ON purchase_order_number.
4. Is the column NOT in this mirror at all (e.g., region head name, customer name, employee, sales rep, line-item quantity from PO)? → SAY SO PLAINLY. Tell the user the closest available alternative if there is one, or that you can't pull it. Don't pretend.

HOW TO BEHAVE:

1. When the user describes a change — acknowledge it briefly, then present a 2-3 line summary of the UPDATED report (what's new, what stayed). Ask:
   Say **"generate"** to apply changes, or tell me what else to adjust.
   Example:
     "I'll add a 'Material name' column by looking up each material_id against the descriptions table. The report stays scoped to the last 30 days of PO line postings. Say **'generate'** to apply, or tell me what else to change."

2. When the user confirms with "generate", "go", "yes", etc. — Return ONLY the FULL updated config JSON. Crucially:
   - INCLUDE the updated SQL in the `sql` field (the runtime will re-run it on preview / download).
   - Update `all_columns` and `columns` to reflect the new SELECT list.
   - Preserve every field the user didn't ask to change (title, filters, group_by, order_by, format).
   Shape:
   {{"ready": true, "config": {{"title": "...", "table": "...", "sql": "SELECT ... FROM `sfml-491907.sap_hana_mirror.<table>` ... LEFT JOIN ... LIMIT 200", "columns": [...], "all_columns": [...], "numeric_columns": [...], "total_columns": [...], "filters": "...", "group_by": "...", "order_by": "...", "aggregations": {{...}}, "format": "excel"}}}}

SQL RULES (apply when you rewrite the `sql`):
- Use fully qualified names: `sfml-491907.sap_hana_mirror.<table>` (with backticks).
- Wrap STRING-YYYYMMDD date columns in `SAFE.PARSE_DATE('%Y%m%d', col)` when filtering. DW fact tables' `posting_date` is already DATE — compare directly.
- For material_documents: each event has 2 rows; filter `debit_credit_indicator='S'` for one-row-per-event aggregation.
- ISSUE/ADJUSTMENT in fact_material_movements_daily are signed — present with natural sign; flip only on explicit "absolute" requests; never `SUM(ABS(...))`.
- LIMIT 200 max on the result set.

CRITICAL RULES:
- The user is editing — they ALREADY know the report context. Don't re-ask things already in current_config (date range, base table, plant scope, etc.).
- ALWAYS return the FULL config (every field), not just the changes. Preserve untouched fields.
- NEVER expose raw column names, table names, SQL, or join syntax to the user in CHAT. Translate to business language ("material name", "plant city", "purchase order date"). The SQL goes ONLY in the JSON config in step 2.
- NEVER output the JSON until the user explicitly confirms with "generate".
- Be concise — max 3-4 sentences per chat message.
"""


REPORT_DESIGN_PROMPT = """You are a tabular report designer for Satori AI, an enterprise intelligence platform connected to the company's SAP ERP mirror.

The user has confirmed a report configuration. Generate ONLY the SQL query needed.

You have access to BigQuery tables in project 'sfml-491907', dataset 'sap_hana_mirror'.
PRIMARY DW REPORTING TABLES (preferred for stock & movement reports — pre-aggregated for performance):
A) fact_material_stock_daily (148M rows) — daily OPENING/CLOSING BALANCE. **3-row-per-key model**: stock_type (STRING — 'STORAGE' / 'SPECIAL' / 'RATE'), plant (STRING), storage_location (STRING — populated only for STORAGE; NULL for SPECIAL/RATE), material_id (STRING), material_type (STRING), base_unit_of_measure (STRING), posting_date (DATE), cumulative_qty (NUMERIC). Per-material qty = SUM(STORAGE.cumulative_qty)+SUM(SPECIAL.cumulative_qty); rate = MAX(RATE.cumulative_qty); value = qty × rate. NO `stock_value` and NO `stock_qty` columns. Date range 2025-04-15 → current. Plants populated: 1100/2100/3100. For "current closing balance" filter to MAX(posting_date); for opening balance pick a specific posting_date. For sloc-filtered queries use the SPECIAL-stock attribution rule (LEFT JOIN a `sm` self-join of materials with STORAGE rows at the target sloc, then count STORAGE qty where sloc matches AND SPECIAL qty where sm.material_id IS NOT NULL).
B) fact_material_movements_daily (651K rows) — daily Receipts / Issues / Adjustments per (plant, storage_location, material_id). Cols: posting_date (DATE), plant, storage_location, material_id, material_type, material_type_flag (STRING — 'H'/'S'), base_unit_of_measure, RECEIPT_QTY, RECEIPT_VALUE, ISSUE_QTY, ISSUE_VALUE, ADJUSTMENT_QTY, ADJUSTMENT_VALUE (all NUMERIC). ISSUE/ADJUSTMENT are mostly NEGATIVE but include positive reversal rows — for net positive magnitude use `-SUM(...)` or `ABS(SUM(...))`, NEVER `SUM(ABS(...))` (double-counts reversals). Date range 2025-05-01 → 2026-02-24.

OTHER TABLES (use fully qualified names `sfml-491907.sap_hana_mirror.<table>`):
1) plants — plant_id, plant_name, plant_name_2, city, country_code, region, purchasing_org, sales_org, distribution_channel, division, valuation_area. Active plant_ids: '1000','1100','1101','2100','3100' (all Lahore, PK).
2) material_master — material_id, material_type (Z113/Z117/Z611/…), base_unit_of_measure (ST/M/EA/KG/…), creation_date (STRING YYYYMMDD), last_change_date.
3) material_descriptions — material_id, language_key ('E'=English (~100%), 'N'=native (2 rows)), material_description. For readable names JOIN on material_id AND language_key='E'.
   MATERIAL ID LEADING-ZERO RULE: stored material_id is 18-char zero-padded (e.g. '000006070000000713'). Users often type the trimmed form (e.g. '6070000000713'). Match with `LTRIM(material_id,'0') = LTRIM(:user_input,'0')` so both forms work; applies to every table that holds material_id.
4) material_valuation — material_id, valuation_area (=plant_id), valuation_type, moving_average_price (NUMERIC, ~5% populated), standard_price (NUMERIC, ~0.4% populated), price_control_indicator ('S'/'V'), price_unit, valuation_class, fiscal_year, fiscal_period. WARNING: total_stock_quantity and total_stock_value are ZERO across all rows — do NOT use them; derive stock value from fact_material_stock_daily as qty × rate per material.
5) material_documents (23.6M rows) — raw movements with movement_type (101 GR / 103/105 blocked GR / 261 GI / 309/310/311/312 sloc transfer / 321 QM release / 344/349 adj / 411/413 transfer / 601 GI delivery / Z87 custom). Cols: material_doc_number, material_id, plant, storage_location, batch, quantity, amount_local_currency, stock_qty, stock_value, debit_credit_indicator ('S'/'H' — 2 rows per event), posting_date (STRING YYYYMMDD), entry_date, document_date, purchase_order, order_number, unit_of_measure. Use ONLY when movement-type-level detail is required.
6) material_documents_date_range — date dimension (date_val DATE).
7) orders — minimal: order_type (Z306/PMWO/CMWO/…), client_id, order_number, creation_date (YYYYMMDD), change_date (YYYYMMDD). NO plant column — derive via material_documents.order_number.
8) purchase_order_header — purchase_order_type (Z001/Z002/NB/ZAN/…), client_id, purchase_order_number, purchase_order_date (YYYYMMDD), change_date (YYYYMMDD). Join accounting_doc_segment on purchase_order_number for amounts + plant.
9) accounting_doc_segment — client_id, company_code ('1000'), belnr, gjahr (STRING '2023'-'2026'), buzei, purchase_order_number, material_id, plant_id, posting_key, debit_credit_indicator ('S'=debit, 'H'=credit), amount_local_currency (PKR), tax_code, posting_date (YYYYMMDD), document_type (ML/WA/WL/WE/RE/SA/KR/KZ/RV/…).
10) universal_journal — client_id, company_code ('1000'), fiscal_year ('2025'/'2026'), doc_number, doc_line, account (GL), transaction_currency (PKR primarily), fiscal_period ('000'-'012'), reference_ledger ('0L'), posting_key, posting_date (YYYYMMDD), material_id, plant_id, amount_local_currency, reference_document.

Notes:
- Numeric columns (cumulative_qty, RECEIPT_*, ISSUE_*, ADJUSTMENT_*, quantity, amount_local_currency, moving_average_price, standard_price) are NUMERIC — aggregate directly (SUM/AVG) without CAST.
- fact_material_stock_daily has NO `stock_value` / `stock_qty` / `rate` columns — those are derived. Per-material qty = SUM(STORAGE.cumulative_qty)+SUM(SPECIAL.cumulative_qty); rate = MAX(RATE.cumulative_qty); value = qty × rate. Aggregate per (plant, material_id) FIRST, then SUM up.
- Do NOT reference material_valuation.total_stock_quantity / total_stock_value (always zero).
- DW fact tables' posting_date is DATE; SAP-style date columns elsewhere (posting_date in material_documents/accounting_doc_segment/universal_journal, purchase_order_date, creation_date, change_date, document_date) are STRING YYYYMMDD — use SAFE.PARSE_DATE('%Y%m%d', <col>) for comparisons.
- For stock on-hand / closing balance reports: use `fact_material_stock_daily` (3-stock_type pattern) filtered to MAX(posting_date). For opening balance reports: pick a specific posting_date. For sloc-filtered balance: SPECIAL stock must be attributed via an `sm` self-join (DISTINCT plant, material_id from STORAGE rows at the target sloc on the date).
- For receipts / issues / adjustments reports: use `fact_material_movements_daily`. ISSUE/ADJUSTMENT are signed (issues typically negative; adjustment_qty often negative but adjustment_value can be positive). Use plain `SUM(...)` and PRESENT WITH NATURAL SIGN — only flip on explicit "absolute" / "magnitude" request. NEVER `SUM(ABS(...))` (double-counts reversals). For sloc-filtered movement reports: keep RECEIPT/ISSUE strict to the sloc; ADJUSTMENT must include the empty-string sloc (`storage_location IN (:sloc, '')`).
- For material_documents, each event has 2 rows (debit_credit_indicator 'S' + 'H'); filter one side (typically 'S') for event counts.
- Active plants IN-list: ('1000','1100','1101','2100','3100'). Stock-fact rows exist only for 1100/2100/3100. Company code: '1000'. Currency: PKR.

Return ONLY valid JSON:
{
  "title": "Report Title",
  "sql": "SELECT ... FROM `sfml-491907.sap_hana_mirror.<table>` ...",
  "columns": ["Column1", "Column2", ...],
  "numeric_columns": ["Column2", ...],
  "total_columns": ["Column2", ...],
  "description": "One-line description"
}

RULES:
- Always use fully qualified table names: `sfml-491907.sap_hana_mirror.TABLE_NAME`
- Column names with spaces MUST be wrapped in backticks in SQL
- Use LIMIT 200 max
- numeric_columns: columns that contain numbers (for right-alignment and number formatting)
- total_columns: columns that should have a SUM total in the footer row
- The SQL must produce a clean tabular result — no nested queries that return prose
- Return ONLY JSON, no markdown fences, no explanation."""


def _extract_response_text(response) -> str | None:
    """
    Robustly extract text from a Gemini response object.
    Tries response.text first, then digs into candidates/parts as fallback.
    """
    # Try the simple path first
    try:
        if response.text:
            return response.text
    except Exception:
        pass

    # Dig into candidates directly
    try:
        candidates = response.candidates
        if candidates and len(candidates) > 0:
            content = candidates[0].content
            if content and content.parts:
                texts = []
                for part in content.parts:
                    if hasattr(part, "text") and part.text:
                        texts.append(part.text)
                if texts:
                    return "\n".join(texts)
    except Exception as e:
        print(f"[REPORT-REFINE] Error extracting from candidates: {e}")

    return None


def _is_confirmation(message: str) -> bool:
    """Check if user message is a confirmation to generate."""
    msg = message.strip().lower()
    confirms = [
        "generate", "go", "go ahead", "yes", "yep", "yeah", "sure",
        "looks good", "do it", "create it", "make it", "proceed",
        "that's good", "perfect", "confirmed", "ok", "okay",
        "let's go", "build it", "run it", "generate it",
    ]
    return msg in confirms or any(msg.startswith(c) for c in confirms)


def refine_report(user_message: str, history: list[dict], existing_config: dict | None = None) -> str:
    """
    Chat-based report refinement. Returns AI response text, or JSON when ready.
    If existing_config is provided, uses edit mode — the AI is shown the
    current config so it can modify the right report instead of inventing
    one from chat history alone.
    """
    client = get_ai_client()

    # Discover tables for context
    tables = discover_tables()
    table_info = []
    for t in tables[:20]:
        schema = get_table_schema(t["full_id"])
        cols = ", ".join(f["name"] for f in schema[:20])
        table_info.append(f"  - {t['table']}: {cols}")
    tables_str = "\n".join(table_info) if table_info else "  (no tables discovered)"

    if existing_config:
        system = REPORT_EDIT_PROMPT.format(
            current_config=json.dumps(existing_config, indent=2),
            tables=tables_str,
        )
    else:
        system = REPORT_REFINE_PROMPT.format(tables=tables_str)

    # Build conversation
    contents = []
    for msg in history:
        contents.append(genai.types.Content(
            role="user" if msg["role"] == "user" else "model",
            parts=[genai.types.Part(text=msg["text"])],
        ))
    contents.append(genai.types.Content(
        role="user",
        parts=[genai.types.Part(text=user_message)],
    ))

    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=contents,
        config=genai.types.GenerateContentConfig(
            system_instruction=system,
            temperature=0.3,
            # 32K matches the dashboard refine. Gemini 2.5 Flash spends
            # internal "thinking" tokens against this budget — earlier 8K
            # was getting eaten by thinking on edit-mode turns where the
            # current_config + table schemas fill the prompt, leaving the
            # visible reply truncated mid-sentence.
            max_output_tokens=32768,
        ),
    )

    # Robustly extract text from the response
    text = _extract_response_text(response)
    if text:
        return text

    # If Gemini returned nothing and user was confirming, try a focused JSON-only call
    print(f"[REPORT-REFINE] response.text was None. Checking if confirmation...")
    if _is_confirmation(user_message) and len(history) >= 2:
        print("[REPORT-REFINE] User confirmed but Gemini returned empty. Retrying with JSON-only prompt...")
        return _retry_generate_config(client, history, tables_str)

    return "I couldn't process that. Could you try again?"


def _retry_generate_config(client, history: list[dict], tables_str: str) -> str:
    """
    Fallback: when the main refine call fails on the confirmation step,
    make a focused call asking Gemini to produce just the JSON config
    based on the conversation so far.
    """
    # Build a summary of what was discussed from the history
    conversation_summary = ""
    for msg in history:
        role = "User" if msg["role"] == "user" else "Satori AI"
        conversation_summary += f"{role}: {msg['text']}\n\n"

    prompt = f"""Based on this report conversation, generate the JSON config to build the report.

CONVERSATION:
{conversation_summary}

AVAILABLE DATA:
{tables_str}

Return ONLY this JSON (no markdown, no explanation, no code fences):
{{"ready": true, "config": {{"title": "Report Title", "table": "BigQueryTableName", "columns": ["col1", "col2"], "filters": "", "group_by": "", "order_by": "", "aggregations": {{"colName": "SUM"}}, "format": "excel"}}}}

Use your intelligence to pick the right table, columns, filters, and aggregations based on what was discussed. Return ONLY the JSON object, nothing else."""

    try:
        retry_response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=genai.types.GenerateContentConfig(
                temperature=0.1,
                # JSON-only path — force the model into structured output mode
                # and give it room to write the full config without truncation.
                max_output_tokens=8192,
                response_mime_type="application/json",
            ),
        )
        retry_text = _extract_response_text(retry_response)
        if retry_text:
            print(f"[REPORT-REFINE] Retry succeeded: {retry_text[:100]}...")
            return retry_text
    except Exception as e:
        print(f"[REPORT-REFINE] Retry also failed: {e}")

    return "I couldn't process that. Could you try again?"


def design_report(user_prompt: str, config: dict = None) -> dict:
    """
    Use Gemini to design a tabular report SQL query.
    If config is provided (from refine flow), use it directly.
    """
    client = get_ai_client()

    tables = discover_tables()
    table_list = ", ".join(t["table"] for t in tables[:30])

    if config:
        # Build from structured config
        prompt = f"""Build a SQL query for this report configuration:
Title: {config.get('title', 'Report')}
Table: {config.get('table', '')}
Columns: {', '.join(config.get('columns', []))}
Filters: {config.get('filters', 'None')}
Group by: {config.get('group_by', 'None')}
Order by: {config.get('order_by', 'None')}
Aggregations: {json.dumps(config.get('aggregations', {}))}

Available tables: {table_list}"""
    else:
        prompt = f"""User request: "{user_prompt}"

Available tables in dataset `sap_hana_mirror`: {table_list}

Design a tabular report query. Focus on producing clean columnar data."""

    # Force JSON output mode + generous token budget. The previous combo
    # (free-form text, 2048 tokens) regularly produced malformed JSON —
    # either truncated mid-SQL or with raw newlines inside string fields,
    # surfacing as "Unterminated string starting at: line 3 column 10".
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt,
        config=genai.types.GenerateContentConfig(
            system_instruction=REPORT_DESIGN_PROMPT,
            temperature=0.3,
            max_output_tokens=16384,
            response_mime_type="application/json",
        ),
    )

    text = (response.text or "").strip()
    if not text:
        raise ValueError("AI returned empty design response.")
    # Belt-and-braces fence stripping — response_mime_type usually prevents
    # fences but the SDK has occasionally let one slip through.
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
    if text.endswith("```"):
        text = text[:-3]
    if text.lstrip().lower().startswith("json"):
        text = text.lstrip()[4:]

    try:
        return json.loads(text.strip())
    except json.JSONDecodeError:
        # Brace-balance fallback: walk the string, find the first `{...}`
        # whose braces balance, and try parsing only that slice. Handles the
        # rare case where the model prepends explanation text before the JSON.
        s = text
        start = s.find("{")
        if start != -1:
            depth, in_str, esc = 0, False, False
            for i in range(start, len(s)):
                c = s[i]
                if esc:
                    esc = False
                    continue
                if c == "\\":
                    esc = True
                    continue
                if c == '"':
                    in_str = not in_str
                    continue
                if in_str:
                    continue
                if c == "{":
                    depth += 1
                elif c == "}":
                    depth -= 1
                    if depth == 0:
                        try:
                            return json.loads(s[start:i + 1])
                        except json.JSONDecodeError:
                            break
        # Last resort: re-raise with a clearer message so the API caller can
        # show "design failed, try again" rather than a raw parser error.
        raise ValueError(
            "AI did not return valid JSON for the report design. "
            "Try rephrasing the request or reducing the column count."
        )


def fetch_report_data(design: dict) -> dict:
    """Execute the SQL query and attach results to the design."""
    sql = design.get("sql", "")
    if sql:
        print(f"[REPORT] Querying: {sql[:120]}...")
        result = run_query(sql, max_rows=200)
        design["data"] = result
    else:
        design["data"] = {"columns": [], "rows": [], "total_rows": 0}
    return design


# ── Excel Generation (Tabular) ──

def generate_excel(design: dict) -> bytes:
    """Generate a clean tabular Excel report with totals row."""
    wb = Workbook()
    ws = wb.active

    title = design.get("title", "Report")
    description = design.get("description", "")
    data = design.get("data", {})
    cols = data.get("columns", [])
    rows = data.get("rows", [])
    numeric_cols = set(design.get("numeric_columns", []))
    total_cols = set(design.get("total_columns", []))

    ws.title = title[:31]
    ws.sheet_properties.tabColor = BRAND_TEAL

    # ── Styles ──
    title_font = Font(name="Calibri", size=14, bold=True, color=BRAND_DARK)
    subtitle_font = Font(name="Calibri", size=10, color="666666")
    meta_font = Font(name="Calibri", size=9, color="999999", italic=True)
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=BRAND_DARK, end_color=BRAND_DARK, fill_type="solid")
    data_font = Font(name="Calibri", size=10, color=BRAND_DARK)
    number_font = Font(name="Calibri", size=10, color=BRAND_TEAL)
    total_font = Font(name="Calibri", size=10, bold=True, color=BRAND_DARK)
    total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    total_label_font = Font(name="Calibri", size=10, bold=True, color=BRAND_DARK)
    zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    thick_top_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="medium", color=BRAND_DARK),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    # ── Title block (rows 1-3) ──
    col_end = get_column_letter(max(len(cols), 1))
    ws.merge_cells(f"A1:{col_end}1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    if description:
        ws.merge_cells(f"A2:{col_end}2")
        ws["A2"] = description
        ws["A2"].font = subtitle_font

    # Header carries the watermark — who exported, when, plus the brand.
    # Audit-friendly: every distributed copy can be traced to a user.
    watermark = design.get("watermark") or ""
    wm_suffix = f" · Exported by {watermark}" if watermark else ""
    ws.merge_cells(f"A3:{col_end}3")
    ws["A3"] = f"Generated: {datetime.datetime.now().strftime('%B %d, %Y %I:%M %p')} · Satori AI · TMC{wm_suffix}"
    ws["A3"].font = meta_font
    ws.row_dimensions[3].height = 18

    # ── Header row (row 5) ──
    header_row = 5
    for c_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=header_row, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="right" if col_name in numeric_cols else "left",
            vertical="center",
        )
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 24

    # ── Data rows (starting row 6) ──
    totals = {c: 0.0 for c in total_cols}
    data_start = header_row + 1

    for r_idx, row_data in enumerate(rows):
        excel_row = data_start + r_idx
        for c_idx, col_name in enumerate(cols, 1):
            raw_val = row_data.get(col_name, "")

            # Convert to number if possible
            val = raw_val
            is_number = False
            if raw_val and raw_val != "None" and raw_val != "null":
                try:
                    val = float(raw_val)
                    is_number = True
                    # Accumulate totals
                    if col_name in total_cols:
                        totals[col_name] += val
                except (ValueError, TypeError):
                    val = str(raw_val) if raw_val != "None" else ""

            cell = ws.cell(row=excel_row, column=c_idx, value=val if val != "None" else "")
            cell.font = number_font if is_number else data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if is_number else "left")

            # Number format for numeric values
            if is_number:
                if abs(val) >= 1000:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0.00'

            # Zebra striping
            if r_idx % 2 == 1:
                cell.fill = zebra_fill

    # ── Totals row ──
    if total_cols and rows:
        total_row = data_start + len(rows)
        ws.row_dimensions[total_row].height = 26

        for c_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=total_row, column=c_idx)
            cell.fill = total_fill
            cell.border = thick_top_border

            if c_idx == 1 and col_name not in total_cols:
                # Label cell
                cell.value = "TOTAL"
                cell.font = total_label_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_name in total_cols:
                cell.value = totals[col_name]
                cell.font = total_font
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if abs(totals[col_name]) >= 1000:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0.00'

    # ── Row count footer ──
    footer_row = data_start + len(rows) + (2 if total_cols else 1)
    ws.merge_cells(f"A{footer_row}:{col_end}{footer_row}")
    total_available = data.get("total_rows", len(rows))
    ws[f"A{footer_row}"] = f"Showing {len(rows)} of {total_available} rows"
    ws[f"A{footer_row}"].font = meta_font

    # ── Auto-width columns ──
    for c_idx, col_name in enumerate(cols, 1):
        sample_lengths = [len(str(col_name))]
        for row_data in rows[:30]:
            sample_lengths.append(len(str(row_data.get(col_name, ""))))
        max_len = max(sample_lengths, default=8)
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 35)

    # ── Freeze panes (freeze header) ──
    ws.freeze_panes = f"A{data_start}"

    # ── Auto-filter ──
    if cols:
        last_col = get_column_letter(len(cols))
        last_data_row = data_start + len(rows) - 1
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"

    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ── PDF Generation (Tabular) ──

def generate_pdf(design: dict) -> bytes:
    """Generate a clean tabular PDF report with totals row."""
    data = design.get("data", {})
    cols = data.get("columns", [])
    rows = data.get("rows", [])
    numeric_cols = set(design.get("numeric_columns", []))
    total_cols = set(design.get("total_columns", []))

    # Use landscape if many columns
    page_size = landscape(A4) if len(cols) > 6 else A4

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=page_size,
        topMargin=15*mm, bottomMargin=15*mm,
        leftMargin=15*mm, rightMargin=15*mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", fontSize=16, fontName="Helvetica-Bold",
                              textColor=colors.HexColor(f"#{BRAND_DARK}"), spaceAfter=4))
    styles.add(ParagraphStyle(name="ReportSubtitle", fontSize=9, fontName="Helvetica",
                              textColor=colors.HexColor("#666666"), spaceAfter=8))
    styles.add(ParagraphStyle(name="ReportMeta", fontSize=7, fontName="Helvetica-Oblique",
                              textColor=colors.HexColor("#999999"), spaceAfter=4))

    elements = []

    # Title
    elements.append(Paragraph(design.get("title", "Report"), styles["ReportTitle"]))
    if design.get("description"):
        elements.append(Paragraph(design["description"], styles["ReportSubtitle"]))
    watermark = design.get("watermark") or ""
    wm_suffix = f" · Exported by {watermark}" if watermark else ""
    elements.append(Paragraph(
        f"Generated: {datetime.datetime.now().strftime('%B %d, %Y %I:%M %p')} · Satori AI · TMC{wm_suffix}",
        styles["ReportMeta"],
    ))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor(f"#{BRAND_GREEN}"), spaceAfter=12))

    if not cols or not rows:
        elements.append(Paragraph("No data available.", styles["ReportSubtitle"]))
    else:
        # ── Build table data ──
        # Format numbers nicely
        def fmt_val(val, col):
            if val is None or val == "None" or val == "null":
                return ""
            if col in numeric_cols:
                try:
                    num = float(val)
                    if abs(num) >= 1000:
                        return f"{num:,.2f}"
                    return f"{num:.2f}"
                except (ValueError, TypeError):
                    pass
            return str(val)[:50]

        table_data = [cols]  # Header row

        # Compute totals while building rows
        totals = {c: 0.0 for c in total_cols}
        for row in rows[:150]:  # Cap for PDF
            table_row = []
            for c in cols:
                raw = row.get(c, "")
                table_row.append(fmt_val(raw, c))
                if c in total_cols:
                    try:
                        totals[c] += float(raw) if raw and raw != "None" else 0
                    except (ValueError, TypeError):
                        pass
            table_data.append(table_row)

        # Totals row
        if total_cols:
            total_row = []
            for i, c in enumerate(cols):
                if i == 0 and c not in total_cols:
                    total_row.append("TOTAL")
                elif c in total_cols:
                    total_row.append(f"{totals[c]:,.2f}" if abs(totals[c]) >= 1000 else f"{totals[c]:.2f}")
                else:
                    total_row.append("")
            table_data.append(total_row)

        # ── Column widths ──
        available_width = page_size[0] - 30*mm
        col_width = available_width / len(cols)
        col_widths = [min(col_width, 140)] * len(cols)

        # ── Style ──
        last_data_row = len(table_data) - 1
        total_row_idx = last_data_row if total_cols else -1

        style_commands = [
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_DARK}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            # Data
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E0E0")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            # Zebra striping
            ("ROWBACKGROUNDS", (0, 1), (-1, total_row_idx - 1 if total_cols else -1),
             [colors.white, colors.HexColor("#F8F9FA")]),
        ]

        # Right-align numeric columns
        for i, c in enumerate(cols):
            if c in numeric_cols:
                style_commands.append(("ALIGN", (i, 0), (i, -1), "RIGHT"))
            else:
                style_commands.append(("ALIGN", (i, 1), (i, -1), "LEFT"))
                style_commands.append(("ALIGN", (i, 0), (i, 0), "LEFT"))

        # Totals row styling
        if total_cols:
            style_commands.extend([
                ("BACKGROUND", (0, total_row_idx), (-1, total_row_idx), colors.HexColor("#E8F5E9")),
                ("FONTNAME", (0, total_row_idx), (-1, total_row_idx), "Helvetica-Bold"),
                ("LINEABOVE", (0, total_row_idx), (-1, total_row_idx), 1.5, colors.HexColor(f"#{BRAND_DARK}")),
            ])

        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle(style_commands))
        elements.append(t)

        # Row count
        elements.append(Spacer(1, 8))
        total_available = data.get("total_rows", len(rows))
        elements.append(Paragraph(
            f"Showing {min(len(rows), 150)} of {total_available} rows",
            styles["ReportMeta"],
        ))

    # Footer
    elements.append(Spacer(1, 16))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor(f"#{BRAND_GREEN}"), spaceAfter=4))
    elements.append(Paragraph("Satori AI · TallyMarks Consulting (TMC)", styles["ReportMeta"]))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


# ── Main orchestrator ──

def generate_report(user_prompt: str, format: str = "excel", config: dict = None, *, watermark: str | None = None) -> tuple[bytes, str, str]:
    """
    Full report generation pipeline.
    Returns: (file_bytes, filename, content_type)
    """
    print(f"[REPORT] Starting report generation for: {user_prompt[:60]}...")

    # Step 1: Design the report SQL.
    # If the caller passed an enriched config (already has `sql` + `all_columns`,
    # e.g. from the preview endpoint or a saved report), skip the AI round-trip
    # and reuse the cached design — keeps download fast and consistent with
    # what the user just previewed.
    cached_sql = ((config or {}).get("sql") or "").strip()
    if cached_sql:
        print(f"[REPORT] Using cached SQL from config ({len(cached_sql)} chars)")
        design = {
            "title": config.get("title", "Report"),
            "description": config.get("description", ""),
            "sql": cached_sql,
            "columns": config.get("all_columns") or config.get("columns") or [],
            "numeric_columns": config.get("numeric_columns", []),
            "total_columns": config.get("total_columns", []),
        }
    else:
        print("[REPORT] Step 1: Designing tabular report with AI...")
        design = design_report(user_prompt, config=config)
        print(f"[REPORT] Design: {design.get('title')}")

    # Step 2: Fetch data from BigQuery
    print("[REPORT] Step 2: Fetching data from BigQuery...")
    design = fetch_report_data(design)

    data = design.get("data", {}) or {}
    # Apply the visible-column subset if the saved config narrowed it. The SQL
    # still selects every column the AI designed, but we prune the rendered
    # output here so a user who clicked "remove" on a column never sees it
    # again, and totals only sum the kept columns.
    visible = (config or {}).get("columns") if config and config.get("columns") else None
    if visible:
        full_cols = data.get("columns", [])
        kept = [c for c in visible if c in full_cols]
        if kept and kept != full_cols:
            print(f"[REPORT] Filtering to {len(kept)}/{len(full_cols)} visible columns")
            data["columns"] = kept
            data["rows"] = [{c: row.get(c) for c in kept} for row in data.get("rows", [])]
            design["data"] = data
            design["columns"] = kept
            design["numeric_columns"] = [c for c in design.get("numeric_columns", []) if c in kept]
            design["total_columns"] = [c for c in design.get("total_columns", []) if c in kept]

    row_count = len(data.get("rows", []))
    print(f"[REPORT] Got {row_count} rows, {len(data.get('columns', []))} columns")

    # Stamp a watermark string on the design so generate_excel/generate_pdf
    # can render it. Lets the file trace back to the user who exported it.
    if watermark:
        design["watermark"] = watermark

    # Step 3: Generate file
    title_slug = design.get("title", "report").replace(" ", "_")[:40]
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")

    if format == "pdf":
        print("[REPORT] Step 3: Generating PDF...")
        file_bytes = generate_pdf(design)
        filename = f"{title_slug}_{timestamp}.pdf"
        content_type = "application/pdf"
    else:
        print("[REPORT] Step 3: Generating Excel...")
        file_bytes = generate_excel(design)
        filename = f"{title_slug}_{timestamp}.xlsx"
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    print(f"[REPORT] Done! Generated {filename} ({len(file_bytes)} bytes)")
    return file_bytes, filename, content_type
