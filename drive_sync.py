#!/usr/bin/env python3
"""
drive_sync.py — upload CSV files from ./data/ to BigQuery.

Workflow:
    1. Drop a CSV into ./data/  (e.g. data/Employee_Data.csv).
    2. Run:    python drive_sync.py
    3. Each CSV becomes a BigQuery table at:
           capability-agent-prod.Satori_Project.<filename-without-.csv>
       Existing tables are REPLACED (WRITE_TRUNCATE) by default; pass
       --append to add rows instead.

Auth:
    Set GOOGLE_APPLICATION_CREDENTIALS to the path of a service-account
    JSON key with `bigquery.dataEditor` + `bigquery.jobUser` on the
    target project. If unset the script falls back to gcloud
    Application Default Credentials.

        # one-time setup, either of these works:
        export GOOGLE_APPLICATION_CREDENTIALS="/path/to/sa-key.json"
        # or
        gcloud auth application-default login

CLI:
    python drive_sync.py                     # upload every CSV in data/
    python drive_sync.py --dry-run           # show what would happen, do nothing
    python drive_sync.py --table Employee_Data
                                             # upload one file: data/Employee_Data.csv
    python drive_sync.py --append            # add rows instead of replacing
    python drive_sync.py --project foo --dataset bar
                                             # override the default target

Filename → table mapping:
    The table name is the CSV filename minus ".csv". Hyphens and spaces
    are converted to underscores so the result is a valid BigQuery
    identifier. So data/Sales Accounts.csv → Sales_Accounts.

Schema:
    Schema is auto-detected from the CSV header + first ~100 rows
    (BigQuery autodetect). If you need an explicit schema, edit
    SCHEMA_OVERRIDES below — keyed by table name.

Notes:
    - All paths are resolved relative to this script's directory so it
      works from any cwd.
    - Files starting with "." or "_" are skipped (lets you keep README
      / template files in data/ without uploading them).
    - Subdirectories under data/ are ignored — keep CSVs flat for now.
"""

from __future__ import annotations

import argparse
import datetime as _dt_top
import os
import re
import sys
import traceback
from pathlib import Path


# ── Tee stdout/stderr to a log file so we have a record even if the
# terminal closes (Windows double-click runs auto-close the window the
# moment the script exits, hiding every error message).
class _Tee:
    """Write to multiple file-like objects at once. Used as a replacement
    for sys.stdout / sys.stderr so each line appears on the terminal AND
    gets persisted to drive_sync.log next to this script."""
    def __init__(self, *streams):
        self._streams = streams
    def write(self, data):
        for s in self._streams:
            try:
                s.write(data)
            except Exception:
                pass
    def flush(self):
        for s in self._streams:
            try:
                s.flush()
            except Exception:
                pass


_LOG_PATH = Path(__file__).resolve().parent / "drive_sync.log"
try:
    _logfile = open(_LOG_PATH, "a", encoding="utf-8", buffering=1)
    _logfile.write("\n" + "=" * 72 + "\n")
    _logfile.write(f"drive_sync run @ {_dt_top.datetime.now().isoformat(timespec='seconds')}\n")
    _logfile.write("argv: " + " ".join(sys.argv) + "\n")
    _logfile.write("=" * 72 + "\n")
    sys.stdout = _Tee(sys.__stdout__, _logfile)
    sys.stderr = _Tee(sys.__stderr__, _logfile)
except Exception as _e:
    print(f"[drive_sync] WARNING: could not open log file ({_e}); continuing without tee")


def _pause_before_exit(code: int) -> None:
    """On Windows, keep the terminal open so the user can read errors.
    Skipped if stdin isn't a TTY (e.g. piped input or CI)."""
    if os.name == "nt" and sys.stdin and sys.stdin.isatty():
        try:
            print("")
            print(f"[drive_sync] finished with exit code {code}. "
                  f"Log saved to {_LOG_PATH}")
            input("Press Enter to close this window...")
        except (EOFError, KeyboardInterrupt):
            pass

# Hard defaults — match the backend's BQ_PROJECT / BQ_DATASET env defaults
# (see backend/main.py line 28-29). These can be overridden per-invocation
# with --project / --dataset or the equivalent env vars.
DEFAULT_PROJECT = os.environ.get("VERTEX_PROJECT", "capability-agent-prod")
DEFAULT_DATASET = os.environ.get("VERTEX_DATASET", "Satori_Project")
DATA_DIR = Path(__file__).resolve().parent / "data"

# Optional per-table schema overrides. By default we let BigQuery infer
# the schema from the CSV. Add an entry here if you want a column forced
# to STRING (e.g. employee codes that look like INTs but should stay
# strings, or dates that need a specific format).
#
# Example:
#     from google.cloud import bigquery
#     SCHEMA_OVERRIDES = {
#         "Employee_Data": [
#             bigquery.SchemaField("Employee_Code", "STRING"),
#             bigquery.SchemaField("Joining_Date", "DATE"),
#         ],
#     }
SCHEMA_OVERRIDES: dict[str, list] = {}


def _safe_table_name(filename: str) -> str:
    """Convert a CSV filename into a valid BigQuery table identifier.

    'Sales Accounts.csv'    -> 'Sales_Accounts'
    'sales-am-scorecard.csv'-> 'sales_am_scorecard'
    'Employee Data v2.csv'  -> 'Employee_Data_v2'
    """
    base = Path(filename).stem
    # BQ table names: letters, digits, underscores. Replace anything else with _.
    safe = re.sub(r"[^A-Za-z0-9_]+", "_", base).strip("_")
    if not safe:
        raise ValueError(f"Could not derive a table name from {filename!r}")
    if safe[0].isdigit():
        safe = "_" + safe
    return safe


def _format_bytes(n: int) -> str:
    """Human-readable byte size."""
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024 or unit == "GB":
            return f"{n:,.1f} {unit}" if unit != "B" else f"{n} {unit}"
        n /= 1024


SUPPORTED_EXTS = (".csv", ".xlsx", ".xlsm")


def _list_csvs(data_dir: Path, only_table: str | None = None) -> list[Path]:
    """Return the sorted list of data files we'll upload (CSV + XLSX).

    Skips:
      - hidden files (starting with '.')
      - underscore-prefixed files (treat as templates / scratch)
      - non-supported extensions (only .csv / .xlsx / .xlsm are accepted)
      - everything under subdirectories of data/
      - lockfiles Excel creates while a workbook is open (e.g. "~$foo.xlsx")

    If --table NAME is given we filter to a single matching file. For
    XLSX workbooks, the match is against the filename-derived base name,
    so multi-sheet workbooks will still expand to N tables at upload time.
    """
    if not data_dir.is_dir():
        return []
    out: list[Path] = []
    for p in sorted(data_dir.iterdir()):
        if not p.is_file():
            continue
        if p.name.startswith((".", "_", "~$")):
            continue
        if p.suffix.lower() not in SUPPORTED_EXTS:
            continue
        if only_table:
            if _safe_table_name(p.name) != only_table:
                continue
        out.append(p)
    return out


def _xlsx_to_csvs(xlsx_path: Path) -> list[tuple[str, Path, "object"]]:
    """Expand a workbook into one temp CSV per visible sheet.

    Returns a list of (table_name, csv_temp_path, tmpdir_handle) tuples.
    Caller is responsible for cleaning up the temp dirs after upload (or
    handing it to the context manager wrapper that does it automatically).

    Hidden + 'veryHidden' sheets are skipped. Single-sheet workbooks
    produce a single table named after the file. Multi-sheet workbooks
    produce N tables, named `{filename_base}__{sheet_name}` so each sheet
    is queryable independently.

    XLSX values are written to the CSV in their formatted-string form:
      - datetimes / dates → ISO 'YYYY-MM-DD' (or 'YYYY-MM-DD HH:MM:SS')
      - numbers           → repr (no scientific notation for ints)
      - None              → empty cell
      - bool              → 'TRUE' / 'FALSE'
      - everything else   → str(value)
    """
    import csv as _csv
    import tempfile
    import datetime as _dt
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit(
            f"Missing dependency to read .xlsx files: {e}\n"
            f"Install with:  pip install openpyxl"
        )

    base = _safe_table_name(xlsx_path.name)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        visible = [name for name in wb.sheetnames
                   if wb[name].sheet_state == "visible"]
        if not visible:
            visible = list(wb.sheetnames)  # fallback if all marked hidden
        out: list[tuple[str, Path, "object"]] = []
        for sheet_name in visible:
            ws = wb[sheet_name]
            if len(visible) == 1:
                table_name = base
            else:
                sheet_safe = re.sub(r"[^A-Za-z0-9_]+", "_", sheet_name).strip("_") or "sheet"
                table_name = f"{base}__{sheet_safe}"

            tmpdir = tempfile.TemporaryDirectory(prefix="drive_sync_")
            csv_path = Path(tmpdir.name) / f"{table_name}.csv"

            def _fmt(v):
                if v is None:
                    return ""
                if isinstance(v, bool):
                    return "TRUE" if v else "FALSE"
                if isinstance(v, _dt.datetime):
                    # Strip the time component if it's midnight (pure date).
                    if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
                        return v.strftime("%Y-%m-%d")
                    return v.strftime("%Y-%m-%d %H:%M:%S")
                if isinstance(v, _dt.date):
                    return v.strftime("%Y-%m-%d")
                if isinstance(v, float) and v.is_integer():
                    # Avoid writing "123.0" for ints loaded from Excel.
                    return str(int(v))
                return str(v)

            with csv_path.open("w", encoding="utf-8", newline="") as fh:
                writer = _csv.writer(fh)
                rows_written = 0
                for row in ws.iter_rows(values_only=True):
                    # Skip wholly-empty rows at the tail (Excel often pads).
                    if rows_written > 0 and all(v is None or v == "" for v in row):
                        continue
                    writer.writerow([_fmt(v) for v in row])
                    rows_written += 1

            out.append((table_name, csv_path, tmpdir))
        return out
    finally:
        wb.close()


def _infer_schema_from_csv(csv_path: Path) -> list:
    """Pre-scan the CSV in Python to infer per-column types and return a
    BigQuery schema. We do this ourselves instead of relying on BigQuery's
    autodetect because BQ samples only the first ~100 rows — so a column
    that starts numeric but later contains values like '20250506-4425'
    gets locked as INT64 and the load aborts ~130k rows in.

    Type detection per column:
      - All values match \\d{4}-\\d{2}-\\d{2}             → DATE
      - All values match integer regex                  → INT64
      - All values match float regex (incl. integers)   → FLOAT64
      - Anything else (any string, mixed, empty)        → STRING

    Empty / all-NULL columns default to STRING (safe).
    """
    import csv as _csv
    from google.cloud import bigquery

    int_re   = re.compile(r"^[+-]?\d+$")
    float_re = re.compile(r"^[+-]?\d+(?:\.\d+)?$|^[+-]?\.\d+$")
    date_re  = re.compile(r"^\d{4}-\d{2}-\d{2}$")

    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = _csv.reader(fh)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"CSV is empty: {csv_path}")
        cols = [h.strip() for h in header]
        # Per-column state: can the column still be each type?
        state = [{"int": True, "float": True, "date": True, "has_value": False}
                 for _ in cols]

        row_count = 0
        for row in reader:
            row_count += 1
            for i, val in enumerate(row[:len(cols)]):
                v = val.strip()
                if not v:
                    continue
                s = state[i]
                s["has_value"] = True
                if s["date"] and not date_re.match(v):
                    s["date"] = False
                if s["int"] and not int_re.match(v):
                    s["int"] = False
                if s["float"] and not float_re.match(v):
                    s["float"] = False

    schema = []
    for col_name, s in zip(cols, state):
        # Sanitize column names so they're valid BQ identifiers.
        safe = re.sub(r"[^A-Za-z0-9_]+", "_", col_name).strip("_") or "col"
        if safe[0].isdigit():
            safe = "_" + safe
        if not s["has_value"]:
            bq_type = "STRING"
        elif s["date"]:
            bq_type = "DATE"
        elif s["int"]:
            bq_type = "INT64"
        elif s["float"]:
            bq_type = "FLOAT64"
        else:
            bq_type = "STRING"
        schema.append(bigquery.SchemaField(safe, bq_type))
    print(f"           inferred schema from {row_count:,} rows: "
          + ", ".join(f"{f.name}={f.field_type}" for f in schema[:6])
          + (f", … +{len(schema)-6} more" if len(schema) > 6 else ""))
    return schema


def _build_load_config(table_name: str, csv_path: Path, append: bool):
    """Construct a BigQuery LoadJobConfig for one CSV."""
    from google.cloud import bigquery
    cfg = bigquery.LoadJobConfig(
        source_format=bigquery.SourceFormat.CSV,
        skip_leading_rows=1,
        write_disposition=(
            bigquery.WriteDisposition.WRITE_APPEND if append
            else bigquery.WriteDisposition.WRITE_TRUNCATE
        ),
        # When appending, the table must already exist — don't try to
        # create it. When replacing, allow creation if missing.
        create_disposition=(
            bigquery.CreateDisposition.CREATE_NEVER if append
            else bigquery.CreateDisposition.CREATE_IF_NEEDED
        ),
        # Be forgiving about quoted commas, embedded newlines, etc.
        allow_quoted_newlines=True,
        allow_jagged_rows=False,
        # Tolerate up to 50 bad rows per CSV before aborting. Raise if
        # you want strict loads; lower if you want zero-tolerance.
        max_bad_records=50,
    )
    if table_name in SCHEMA_OVERRIDES:
        cfg.schema = SCHEMA_OVERRIDES[table_name]
        cfg.autodetect = False
    else:
        # Pre-scan the CSV ourselves and supply an explicit schema —
        # bypasses BigQuery's tiny-sample autodetect that misclassifies
        # ID-shaped columns containing hyphens (e.g. '20250506-4425').
        try:
            cfg.schema = _infer_schema_from_csv(csv_path)
            cfg.autodetect = False
        except Exception as e:
            print(f"           local schema inference failed ({e}); "
                  f"falling back to BigQuery autodetect")
            cfg.autodetect = True
    return cfg


def _get_client(project: str):
    """Build an authenticated BigQuery client.

    Mirrors the auth path in backend/bigquery_client.py: prefer an
    explicit service-account key (GOOGLE_APPLICATION_CREDENTIALS) and
    fall back to gcloud Application Default Credentials.
    """
    from google.cloud import bigquery

    sa_key = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    if sa_key:
        if not Path(sa_key).is_file():
            raise SystemExit(
                f"GOOGLE_APPLICATION_CREDENTIALS points to {sa_key!r} "
                f"but that file does not exist."
            )
        from google.oauth2 import service_account
        creds = service_account.Credentials.from_service_account_file(sa_key)
        return bigquery.Client(project=project, credentials=creds)

    # ADC — works after `gcloud auth application-default login`.
    try:
        return bigquery.Client(project=project)
    except Exception as e:
        raise SystemExit(
            "Could not authenticate to BigQuery.\n"
            "Either set GOOGLE_APPLICATION_CREDENTIALS to a service-account "
            "JSON key path, or run `gcloud auth application-default login`.\n"
            f"Underlying error: {e}"
        )


def _ensure_dataset(client, project: str, dataset: str) -> None:
    """Create the dataset if it doesn't exist. Region defaults to US so
    it matches the existing Satori_Project location; change if your
    capability-agent-prod data lives elsewhere."""
    from google.cloud import bigquery
    from google.api_core.exceptions import NotFound
    ds_ref = f"{project}.{dataset}"
    try:
        client.get_dataset(ds_ref)
    except NotFound:
        ds = bigquery.Dataset(ds_ref)
        ds.location = os.environ.get("BQ_LOCATION", "US")
        client.create_dataset(ds, exists_ok=True)
        print(f"[drive_sync] created dataset {ds_ref} (location={ds.location})")


def upload_one(client, csv_path, project, dataset,
               append, dry_run,
               table_name=None,
               display_name=None) -> dict:
    """Upload a single CSV -> BigQuery table. Returns a summary dict.

    `table_name` defaults to the safe-name of the CSV filename, but the
    XLSX expansion path passes an explicit one (so multi-sheet workbooks
    can land at `Filename__SheetName`). `display_name` is what we print
    in logs.
    """
    if table_name is None:
        table_name = _safe_table_name(csv_path.name)
    if display_name is None:
        display_name = csv_path.name
    table_ref = f"{project}.{dataset}.{table_name}"
    size = csv_path.stat().st_size
    mode = "APPEND" if append else "REPLACE"
    print(f"[drive_sync] {display_name}  ->  {table_ref}  "
          f"({_format_bytes(size)}, mode={mode})")

    if dry_run:
        return {"table": table_name, "rows": None, "skipped": "dry-run"}

    cfg = _build_load_config(table_name, csv_path=csv_path, append=append)
    with csv_path.open("rb") as fh:
        job = client.load_table_from_file(fh, table_ref, job_config=cfg)
    job.result()

    table = client.get_table(table_ref)
    rows = table.num_rows
    print(f"           done -- {rows:,} rows, {len(table.schema)} columns")
    return {"table": table_name, "rows": rows, "columns": len(table.schema)}


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Upload CSVs / XLSX workbooks from ./data/ to BigQuery.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--project", default=DEFAULT_PROJECT,
                        help=f"GCP project (default: {DEFAULT_PROJECT})")
    parser.add_argument("--dataset", default=DEFAULT_DATASET,
                        help=f"BigQuery dataset (default: {DEFAULT_DATASET})")
    parser.add_argument("--table", default=None,
                        help="Upload only the file that maps to this table name "
                             "(e.g. --table Employee_Data -> data/Employee_Data.csv).")
    parser.add_argument("--append", action="store_true",
                        help="Append rows to existing tables instead of replacing them.")
    parser.add_argument("--dry-run", action="store_true",
                        help="List what would be uploaded but make no changes.")
    parser.add_argument("--data-dir", default=str(DATA_DIR),
                        help=f"Directory of files to upload (default: {DATA_DIR})")
    args = parser.parse_args()

    data_dir = Path(args.data_dir).resolve()
    csvs = _list_csvs(data_dir, only_table=args.table)
    if not csvs:
        if args.table:
            print(f"[drive_sync] no file in {data_dir} maps to table {args.table!r}.")
        else:
            print(f"[drive_sync] no .csv / .xlsx files found in {data_dir}.")
            print(f"[drive_sync] Drop your files in {data_dir} and re-run.")
        return 0

    print(f"[drive_sync] target: {args.project}.{args.dataset}")
    print(f"[drive_sync] source: {data_dir}  ({len(csvs)} file{'s' if len(csvs) != 1 else ''})")
    if args.dry_run:
        print(f"[drive_sync] DRY RUN -- no uploads will happen.")

    client = None
    if not args.dry_run:
        client = _get_client(args.project)
        _ensure_dataset(client, args.project, args.dataset)

    results = []
    errors = []
    for data_file in csvs:
        ext = data_file.suffix.lower()
        if ext in (".xlsx", ".xlsm"):
            if args.dry_run:
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(data_file, read_only=True, data_only=True)
                    visible = [n for n in wb.sheetnames if wb[n].sheet_state == "visible"] or list(wb.sheetnames)
                    base = _safe_table_name(data_file.name)
                    for sheet_name in visible:
                        if len(visible) == 1:
                            tn = base
                        else:
                            ss = re.sub(r"[^A-Za-z0-9_]+", "_", sheet_name).strip("_") or "sheet"
                            tn = f"{base}__{ss}"
                        size = data_file.stat().st_size
                        mode = "APPEND" if args.append else "REPLACE"
                        print(f"[drive_sync] {data_file.name} [{sheet_name}]  ->  "
                              f"{args.project}.{args.dataset}.{tn}  "
                              f"({_format_bytes(size)}, mode={mode})")
                        results.append({"table": tn, "rows": None, "skipped": "dry-run"})
                    wb.close()
                except Exception as e:
                    print(f"           FAILED -- {e}")
                    errors.append((data_file.name, str(e)))
                continue
            try:
                tasks = _xlsx_to_csvs(data_file)
            except SystemExit:
                raise
            except Exception as e:
                print(f"[drive_sync] {data_file.name}: failed to read workbook: {e}")
                errors.append((data_file.name, str(e)))
                continue
        else:
            tasks = [(_safe_table_name(data_file.name), data_file, None)]

        for table_name, csv_path, cleanup in tasks:
            display = data_file.name if cleanup is None else f"{data_file.name} [{csv_path.stem}]"
            try:
                r = upload_one(
                    client=client,
                    csv_path=csv_path,
                    table_name=table_name,
                    display_name=display,
                    project=args.project,
                    dataset=args.dataset,
                    append=args.append,
                    dry_run=args.dry_run,
                )
                results.append(r)
            except Exception as e:
                print(f"           FAILED -- {e}")
                errors.append((display, str(e)))
            finally:
                if cleanup is not None:
                    try:
                        cleanup.cleanup()
                    except Exception:
                        pass

    print()
    print(f"[drive_sync] summary: {len(results)} ok, {len(errors)} failed")
    if errors:
        for name, msg in errors:
            print(f"   ! {name}: {msg}")
        return 1
    return 0


def _preflight() -> None:
    """Validate dependencies + auth before doing any work, with friendly
    messages instead of cryptic ImportError / AuthError mid-run."""
    missing = []
    try:
        import google.cloud.bigquery  # noqa: F401
    except ImportError:
        missing.append("google-cloud-bigquery")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        missing.append("openpyxl")
    if missing:
        print("[drive_sync] FATAL - missing Python packages: " + ", ".join(missing))
        print(f"[drive_sync] Install with: pip install {' '.join(missing)}")
        raise SystemExit(2)

    sa_key = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    if sa_key:
        if not Path(sa_key).is_file():
            print(f"[drive_sync] FATAL - GOOGLE_APPLICATION_CREDENTIALS "
                  f"points to {sa_key!r} but that file does not exist.")
            raise SystemExit(2)
        print(f"[drive_sync] using service-account key: {sa_key}")
    else:
        try:
            import google.auth
            creds, project = google.auth.default()
            print(f"[drive_sync] using Application Default Credentials "
                  f"(project: {project or 'unset'})")
        except Exception as e:
            print("[drive_sync] FATAL - no credentials configured.")
            print("[drive_sync] Either:")
            print("[drive_sync]   1. Set GOOGLE_APPLICATION_CREDENTIALS to a "
                  "service-account JSON key path, OR")
            print("[drive_sync]   2. Run:  gcloud auth application-default login")
            print(f"[drive_sync] Underlying error: {e}")
            raise SystemExit(2)


if __name__ == "__main__":
    exit_code = 1
    try:
        _preflight()
        exit_code = main()
    except SystemExit as e:
        exit_code = e.code if isinstance(e.code, int) else (0 if e.code is None else 1)
    except KeyboardInterrupt:
        print("\n[drive_sync] interrupted by Ctrl+C")
        exit_code = 130
    except Exception as e:
        print("\n[drive_sync] UNCAUGHT EXCEPTION:")
        print(traceback.format_exc())
        print(f"[drive_sync] {type(e).__name__}: {e}")
        exit_code = 1
    finally:
        _pause_before_exit(exit_code)
    sys.exit(exit_code)
